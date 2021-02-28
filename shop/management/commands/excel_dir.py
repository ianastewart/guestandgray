import os
from PIL import Image
from django.core.management.base import BaseCommand
from openpyxl import Workbook

from shop.management.utils import sanitize


class Command(BaseCommand):
    help = "Create excel file containing filenames"

    def add_arguments(self, parser):
        parser.add_argument("directory", type=str, help="Source directory path")

    def handle(self, *args, **options):
        path = options["directory"]
        dir_name = "Excel directory.xlsx"
        if not os.path.exists(path):
            self.stdout.write(f"{path} does not exist")
            return
        target = os.path.join(path, dir_name)
        try:
            os.remove(target)
        except OSError as e:
            if e.errno != 2:
                self.stdout.write(e.strerror)
                return
        book = Workbook()
        sheet = book.active
        files = os.listdir(path)
        files.sort()
        row = 1
        count = 0
        print(f"Processing {len(files)} images")
        for name in files:
            if name != dir_name:
                sheet.cell(row=row, column=1).value = name
                e = None
                try:
                    ref, changed, new_name = sanitize(name)
                except Exception as e:
                    print(name, str(e))
                if e is None:
                    sheet.cell(row=row, column=2).value = ref
                    if changed:
                        sheet.cell(row=row, column=3).value = new_name
                    file_path = os.path.join(path, name)
                    try:
                        with Image.open(file_path) as im:
                            sheet.cell(row=row, column=4).value = os.path.getsize(
                                file_path
                            )
                            sheet.cell(row=row, column=5).value = im.width
                            sheet.cell(row=row, column=6).value = im.height
                            if im.height == im.width:
                                shape = "square"
                            elif im.height > im.width:
                                shape = "portait"
                            else:
                                shape = "landscape"
                            sheet.cell(row=row, column=7).value = shape
                    except Exception as e:
                        print(name, str(e))
                    row += 1
                    count += 1
                    if count == 100:
                        print(f"{row} processed")
                        count = 0
        book.save(target)
        print("Done")
