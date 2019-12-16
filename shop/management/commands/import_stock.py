import os
import datetime
from decimal import *
from django.conf import settings
from django.core.management.base import BaseCommand
from shop.models import Item, Contact, Address, Invoice, Purchase, Lot
from shop.truncater import truncate
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Process Excel stock file"

    def add_arguments(selfself, parser):
        parser.add_argument("workbook", type=str, help="workbook name")

    def handle(self, *args, **options):
        path = os.path.join(settings.MEDIA_ROOT, "excel", options["workbook"])
        if not os.path.exists(path):
            self.stdout.write(f"{path} does not exist")
        else:
            try:
                self.stdout.write(f">> Start load workbook")
                wb = load_workbook(filename=path, read_only=True, data_only=True)
                self.stdout.write(f"Workbook loaded")
                ws = wb.get_sheet_by_name("Buyers")
                self.import_buyers(ws)
                ws = wb.get_sheet_by_name("Vendors")
                self.import_vendors(ws)
                for ws in wb.worksheets:
                    if "Stock" in ws.title:
                        self.import_stock(ws)
            except Exception as e:
                self.stdout.write(f"Exception processing sheet {ws.title} \n{str(e)}")

    def import_buyers(self, ws):
        """ Import buyers and create invoices """
        self.stdout.write(">> Start import buyers")
        try:
            rowgen = ws.rows
            cols = [c.value for c in next(rowgen)]
            try:
                col_date = cols.index("Date")
                col_invoice_no = cols.index("Invoice no")
                col_address = cols.index("Address")
            except ValueError as e:
                self.stdout.write(f"Header error in sheet {ws.title}:\n{str(e)}")
                return

            Contact.objects.all().delete()
            Invoice.objects.all().delete()
            row_number = 0
            new_contacts = 0
            existing_contacts = 0
            new_invoices = 0
            errors = 0
            for row in rowgen:
                dat = row[col_date].value
                number = row[col_invoice_no].value
                adr = row[col_address].value
                buyer = None
                if adr:
                    row_number = row[col_address].row
                elif number:
                    row_number = row[col_invoice_no].row
                # if row_number == 8:
                #     breakpoint()

                if adr:
                    first_name, name, address = parse_name_address(adr)
                    contacts = Contact.objects.filter(company=name)
                    if len(contacts) > 0:
                        for contact in contacts:
                            if contact.main_address.address[:10] == address[:10]:
                                buyer = contact
                                existing_contacts += 1
                                if not buyer.buyer:
                                    buyer.buyer = True
                                    buyer.save()
                                break
                if not buyer and adr:

                    buyer = Contact.objects.create(
                        first_name=first_name, company=name, buyer=True
                    )
                    new_contacts += 1
                    buyer_address = Address.objects.create(
                        address=address, contact=buyer
                    )
                    buyer.main_address = buyer_address
                    buyer.save()
                if number:
                    date = parse_date(dat)  # date defaults if missing
                    try:
                        Invoice.objects.create(
                            date=date,
                            number=number,
                            buyer=buyer,
                            address=buyer_address,
                            total=0,
                            paid=True,
                        )
                        new_invoices += 1
                    except Exception as e:
                        self.stdout.write(
                            f"Row {row_number} Duplicate invoice: {number} ignored"
                        )
                        errors += 1
            self.stdout.write(
                f"New contacts: {new_contacts}, Existing contacts: {existing_contacts}, Invoices: {new_invoices} Errors: {errors}"
            )
        except Exception as e:
            self.stdout.write(f"Exception in {ws.title} row {row_number}\n{str(e)}")
            raise

    def import_vendors(self, ws):
        try:
            self.stdout.write(">> Start import vendors")
            Contact.objects.filter(vendor=True).delete()
            count = 0
            done = False
            for row in ws.iter_rows(min_row=1, min_col=2, max_row=500, max_col=2):
                for cell in row:
                    if cell.value:
                        name, address = parse_name_address(cell.value, vendor=True)
                        notes = vendor_key(cell.value)
                        try:
                            Contact.objects.get(
                                company=name, main_address__address=address, vendor=True
                            )
                        except Contact.DoesNotExist:
                            contact = Contact.objects.create(
                                company=name, notes=notes, vendor=True
                            )
                            address = Address.objects.create(
                                address=address, contact=contact
                            )
                            contact.main_address = address
                            contact.save()
                            count += 1
                    else:
                        done = True
                if done:
                    break
        except Exception as e:
            self.stdout.write("Exception in vendors")
        self.stdout.write(f"End import vendors: {count} contacts created")

    def import_stock(self, ws):
        """ Read stock sheet, create missing items, update purchases and invoices """
        self.stdout.write(">> Start import stock")
        Item.objects.filter(category__isnull=True).delete()
        Purchase.objects.all().delete()
        rowgen = ws.rows
        cols = [c.value for c in next(rowgen)]
        try:
            col_pdate = cols.index("PDate")
            col_vendor = cols.index("Vendor")
            col_vat = cols.index("VAT")
            col_lot = cols.index("Lot No.")
            col_prem = cols.index("Prem")
            col_cost_lot = cols.index("CostLot")
            col_cost_item = cols.index("CostItem")
            col_cost_rest = cols.index("Cost Rest")
            col_stock_no = cols.index("Stock No")
            col_description = cols.index("Full Description")
            col_price = cols.index("Price")
            col_sale_date = cols.index("SaleDate")
            col_inv_no = cols.index("InvNo")
            col_section = cols.index("Section Text")
        except ValueError as e:
            self.stdout.write(f"Header error in sheet {ws.title}:\n{str(e)}")
            raise
        exists = 0
        created = 0
        vendors_created = 0
        inv_created = 0
        last_vendor = None
        last_pdate = None
        try:
            for row in rowgen:
                row_number = row[0].row
                ref = row[col_stock_no].value
                description = row[col_description].value
                price, text = parse_decimal(row[col_price].value)
                if text:
                    self.stdout.write(f"Ignoring row: {row[0].row} Price = {text}")
                    continue
                inv_no = row[col_inv_no].value
                sale_date = row[col_sale_date].value
                archive = False
                if inv_no and sale_date:
                    archive = True
                    sale_date = parse_date(sale_date)
                vat, _ = parse_decimal(row[col_vat].value)
                margin_scheme = not vat
                lot_number = row[col_lot].value
                cost_item, _ = parse_decimal(row[col_cost_item].value)
                cost_lot, _ = parse_decimal(row[col_cost_lot].value)
                cost_rest, _ = parse_decimal(row[col_cost_rest].value)
                premium, _ = parse_decimal(row[col_prem].value)
                pdate = parse_date(row[col_pdate].value)
                vendor_name = parse_special(row[col_vendor].value)
                # Find or create a vendor
                vendor = None
                if vendor_name:
                    key = vendor_key(vendor_name)
                    vendors = Contact.objects.filter(notes=key)
                    l = len(vendors)
                    if l == 0:
                        name, address = parse_name_address(vendor_name, vendor=True)
                        vendor = Contact.objects.create(
                            company=name, notes=key, vendor=True
                        )
                        address = Address.objects.create(
                            address=address, contact=vendor
                        )
                        vendor.main_address = address
                        vendor.save()
                        vendors_created += 1
                    elif l == 1:
                        vendor = vendors[0]
                    else:
                        self.stdout.write(
                            f"Info row: {row_number} Multiple vendors {l}: {vendor}"
                        )
                else:
                    # Missing vendor uses previous vendor if same purchase date
                    vendor = last_vendor
                    # if last_pdate == pdate:
                    # self.stdout.write(
                    #     f"Info row: {row_number} {vendor.name} different date"
                    # )

                # Find or create an item and add costs
                try:
                    item = Item.objects.get(ref=ref)
                    if item.sale_price != price:
                        if not item.archive:
                            self.stdout.write(
                                f"Info row: {row_number} Web price: {item.sale_price} Excel price: {price}"
                            )
                    exists += 1
                except Item.DoesNotExist:
                    item = Item.objects.create(
                        name=truncate(description),
                        ref=ref,
                        description=description,
                        sale_price=Decimal(price),
                        category=None,
                        image=None,
                        archive=archive,
                        visible=False,
                    )
                    created += 1
                item.cost_price = cost_item
                item.restoration_cost = cost_rest

                # Create a purchase record and link item to it
                try:
                    purchase = Purchase.objects.get(vendor=vendor, date=pdate)
                except Purchase.DoesNotExist:
                    purchase = Purchase.objects.create(
                        date=pdate,
                        invoice_number=0,
                        invoice_total=0,
                        buyers_premium=premium,
                        vendor=vendor,
                        margin_scheme=margin_scheme,
                        vat=vat,
                    )
                # Find or create a lot
                try:
                    lot = Lot.objects.get(purchase=purchase, number=lot_number)
                except Lot.DoesNotExist:
                    lot = Lot.objects.create(
                        purchase=purchase, number=lot_number, cost=cost_lot
                    )
                item.lot = lot
                # if sold, update invoice with item
                if inv_no:
                    try:
                        invoice = Invoice.objects.get(number=inv_no)
                        item.invoice = invoice
                    except Invoice.DoesNotExist:
                        invoice = Invoice(
                            date=pdate, number=inv_no, buyer=None, total=0, paid=True
                        )
                        inv_created += 1
                    item.invoice = invoice
                    invoice.total += item.sale_price
                    invoice.save()
                item.save()
                last_pdate = pdate
                last_vendor = vendor

        except Exception as e:
            self.stdout.write(f"Exception in {ws.title} row {row_number}\n{str(e)}")
            raise
        # Calculate the invoice total for every purchase
        for purchase in Purchase.objects.all():
            total = purchase.buyers_premium
            for lot in purchase.lot_set.all():
                total += lot.cost
            purchase.invoice_total = total
            purchase.save()

        self.stdout.write(
            f"Items exists: {exists} Created: {created} Vendors created {vendors_created} Invoices created: {inv_created}"
        )


def parse_name_address(value, vendor=False):
    """ Separate name and address by first comma or first numeral """

    try:
        value = despace(value).replace("'", "")
        value = parse_special(value)
        first_name = ""
        name = ""
        address = ""
        comma = value.find(",")
        if "Ltd" in value:
            i = value.find("Ltd")
            if i > 0:
                i += 3
                if i == len(value):
                    name = value
                else:
                    if value[i + 1] == ".":
                        i += 1
                    name = value[: i + 1].strip()
                    address = value[i + 2 :]
            else:
                name = value
        elif comma > 0:
            name = value[:comma].strip()
            address = value[comma + 1 :]
        else:
            i = has_digit(value)
            if i and i < 40 and value[i - 1] == " " or value[i - 1] == ",":
                name = value[:i].strip()
                address = value[i:]
            elif "&" in value:
                i = value.find("&")
                j = value[i + 2 :].find(" ")
                if j > 0:
                    name = value[: j + i + 2]
                    address = value[j + i + 3 :]
                else:
                    name = value
            else:
                space = value.find(" ")
                if space > 0 and space < 4:
                    s = value[space + 1 :].find(" ")
                    if s > 0:
                        space = space + 1 + s
                if space > 0:
                    name = value[:space].strip()
                    address = value[space + 1 :]
                else:
                    name = value
        address = address.strip()
        if vendor:
            address = address.replace(", ", "\n").replace(",", "\n")
            return name, address
        # Try for a first name
        save_address = address
        if len(address) > 0 and not address[0].isdigit():
            comma = address.find(",")
            if comma > 0 and comma < 30:
                first_name = address[:comma].strip()
                address = address[comma + 1 :]
            else:
                at = address.find("@")
                if at > 0:
                    if "Anne" in address:
                        first_name = "Anne"
                        address = address[at + 1 :]
                else:
                    i = has_digit(address)
                    if i < 20:
                        first_name = address[:i].strip()
                        address = address[i:]
                    else:
                        space = address.find(" ")
                        if space > 0 and space < 30:
                            first_name = address[:space].strip()
                            address = address[space + 1 :]
        address = address.strip().replace(", ", "\n").replace(",", "\n")
        if has_digit(first_name):  # or len(first_name) > 30:
            address = save_address
            first_name = ""
        return first_name, name, address
    except Exception as e:
        raise


def parse_special(value):
    if value:
        if value == "24 covers":
            return ""
        if value[:7] == "Bonhams":
            if "Knowle" in value:
                value = value.replace("Knowle", "Knowle,").replace("Rd", "Rd,")
            elif "Montpelier St " in value:
                value = value.replace("Bonhams ", "Bonhams, ").replace("St ", "St, ")
            elif "Oxford" in value:
                value = value.replace("Bonhams, Ox", "Bonhams Ox")
            elif "Dowell" in value:
                value = "Bonhams West Country, Dowell Street, Honiton, Devon EX14 8LX"
        elif "Rasmussen Bredgade" in value:
            value = "Rasmussen Bredgade 33 1260 Copenhagen K Denmark"
        elif "Bukowski" in value:
            value = "Lilla Bukowskis, Skomakaregatan 12 211 34, Malmo, Sweden"
        elif "AB Stockholm" in value:
            value = "AB Stockholms Auktionsverk, Jacobsgatan 10, Box 16256, 103 25 Stockholm, Sweden"
    return value


def has_digit(string):
    """ return 0 if no digit else position of first digit """
    for i, c in enumerate(string):
        if c.isdigit():
            return i
    return 0


def parse_date(value):
    """ Parse date in form dd.mm.yy """
    if value:
        if type(value) is datetime.datetime:
            return value
        value = value.replace("..", ".").replace(",", ".").replace(" ", ".")
        if value == "26.3-3.4.07":
            value = "3.4.7"
        if value == ".8.19":
            value = "1.8.19"
        bits = value.split(".")
        if len(bits) >= 3:
            # dates with day range use end date
            i = bits[0].find("-")
            if i > 0:
                bits[0] = bits[0][i + 1 :]
            i = bits[0].find("/")
            if i > 0:
                bits[0] = bits[0][i + 1 :]
            i = bits[0].find("&")
            if i > 0:
                bits[0] = bits[0][i + 1 :]
            day = int(bits[0])
            month = int(bits[1])
            year = int(bits[2])
            # handle known rogue year values
            if year > 100:
                year -= 200
            if year > 50:
                year += 1900
            else:
                year += 2000
            if month == 35:
                month = 5
            # February 31!
            if day == 31 and month == 2:
                day = 28
            return datetime.date(year, month, day)
        elif value == "O23.85L":
            return datetime.date(2015, 9, 21)
        else:
            value = value
    return datetime.date(1900, 1, 1)


def parse_decimal(value):
    if type(value) is int:
        return Decimal(value), ""
    elif type(value) is float:
        return Decimal(value), ""
    elif type is str:
        if value:
            return Decimal(0), value
    return Decimal(0), ""


def vendor_key(value):
    return value.strip().replace(" ", "").replace(",", "").replace("-", "")[:30]


def despace(value):
    while "  " in value:
        value = value.replace("  ", " ")
    return value
